
import traceback
from conf.logconfig import logger
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill

class ExcelUtils():


    def __init__(self):
        logger.info('--')


    def makeTableList(self, tables):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "테이블 정의서"

            ws2 = wb.create_sheet(title="테이블 목록")

            sheet2 = []
            sheet2.append(['No','Table', 'Table Name'])
            for i, item in enumerate(tables):
                sheet2.append([i,item['TABLE_NAME'], item['COMMENTS']])
            # 데이터 엑셀 시트에 쓰기
            for row in sheet2:
                ws2.append(row)
            
            # 데이터 설정
            data = []
            rowGrpCnts = []
            rowCnt = 1
            for i, item in enumerate(tables):

                areaCnt = 0
                colCnt = 0
                indexCnt = 0
                data.append(['Table Name', item['TABLE_NAME'], 'SYNONYM',f'{item["OWNER"]}.{item["TABLE_NAME"]}'])
                data.append(['Table Description', item['COMMENTS']])
                data.append(["Column Name", "Description", "Type", "Size", "ID", "PK", "FK", "NULL", "Default", "비고"])
                data.append([])
                areaCnt += 4
                if item.get('COL_INFO') is not None:
                    colunms = item['COL_INFO']
                    areaCnt += len(colunms)
                    colCnt = len(colunms)
                    for j, item2 in enumerate(colunms):
                        data.append([item2['COLUMN_NAME'], item2['COMMENTS'], item2['DATA_TYPE'],item2['DATA_LENGTH'],item2['COLUMN_ID'],item2['IS_PRIMARY_KEY'],item2['IS_FOREIGN_KEY'],item2['NULLABLE'],'',''])

                data.append(["업무규칙"]),
                data.append(['INDEX',"Index Name", "Index Column", "Sort", "", "No"]),
                areaCnt += 2
                if item.get('TABLE_INDEX') is not None:
                    colunms2 = item['TABLE_INDEX']
                    areaCnt += len(colunms2)
                    indexCnt = len(colunms2)
                    for z, item3 in enumerate(colunms2):
                        data.append(['INDEX',item3['INDEX_NAME'], item3['COLUMN_NAME'],item3['DESCEND'],'', item3['COLUMN_POSITION']])

                data.append([])
                data.append([])
                data.append([])
                areaCnt += 3 # 그룹별
                rowGrpCnts.append((rowCnt, areaCnt,colCnt,indexCnt))
                rowCnt += areaCnt

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            top_thick_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thick'),
                                     bottom=Side(style='thin'))

            bom_thick_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thick'))

            header_fill = PatternFill(start_color="b1b0b0", end_color="b1b0b0", fill_type="solid")


            # 데이터 엑셀 시트에 쓰기
            for row in data:
                ws.append(row)


            for start, cnt , colcnt, indexcnt in rowGrpCnts:
                print(f'start:{start}, cnt:{cnt}, colcnt:{colcnt}, indexcnt:{indexcnt}')
                # 테이블명
                ws.merge_cells(start_row=start, start_column=4, end_row=start, end_column=10)
                # 테이블 desciprt 병합
                ws.merge_cells(start_row=start+1, start_column=2, end_row=start+1, end_column=10)
                # 칼럼해더 병합
                ws.merge_cells(start_row=start + 2, start_column=1, end_row=start + 3, end_column=1)
                ws.merge_cells(start_row=start + 2, start_column=2, end_row=start + 3, end_column=2)
                ws.merge_cells(start_row=start + 2, start_column=3, end_row=start + 3, end_column=3)
                ws.merge_cells(start_row=start + 2, start_column=4, end_row=start + 3, end_column=4)
                ws.merge_cells(start_row=start + 2, start_column=5, end_row=start + 3, end_column=5)
                ws.merge_cells(start_row=start + 2, start_column=6, end_row=start + 3, end_column=6)
                ws.merge_cells(start_row=start + 2, start_column=7, end_row=start + 3, end_column=7)
                ws.merge_cells(start_row=start + 2, start_column=8, end_row=start + 3, end_column=8)
                ws.merge_cells(start_row=start + 2, start_column=9, end_row=start + 3, end_column=9)
                ws.merge_cells(start_row=start + 2, start_column=10, end_row=start + 3, end_column=10)
                ws.merge_cells(start_row=start + 2 + colcnt + 2, start_column=2, end_row=start + 2 + colcnt + 2, end_column=10)
                ws.merge_cells(start_row=start + 2 + colcnt + 3, start_column=1, end_row=start + 2 + colcnt + 3+indexcnt, end_column=1)


                next = (start + 2 + colcnt + 2)
                for i in range(indexcnt+1):
                    nextIndex = next+(i+1)
                    ws.merge_cells(start_row=nextIndex, start_column=4, end_row=nextIndex, end_column=5)
                    ws.merge_cells(start_row=nextIndex, start_column=6, end_row=nextIndex, end_column=10)



            # 스타일 적용 (1행부터 10행, 1열부터 5열)
            for start, cnt , colcnt, indexcnt in rowGrpCnts:
                for row in ws.iter_rows(min_row=start, max_row=start+(cnt-4), min_col=1, max_col=10):
                        for cell in row:
                            cell.border = thin_border

                for row in ws.iter_rows(min_row=start, max_row=start, min_col=1, max_col=10):
                        for cell in row:
                            cell.border = top_thick_border

                for row in ws.iter_rows(min_row=start+(cnt-4), max_row=start+(cnt-4), min_col=1, max_col=10):
                        for cell in row:
                            cell.border = bom_thick_border

            # bold_font = Font(bold=True, color="eee9e9")
            bold_font = Font(bold=True)
            for row in ws.iter_rows():
                if row[0].value == 'Table Name':
                    row[0].font = bold_font
                    row[0].fill = header_fill
                    row[2].font = bold_font
                    row[2].fill = header_fill
                if row[0].value == 'Table Description':
                    row[0].font = bold_font
                    row[0].fill = header_fill
                if row[0].value == 'Column Name':
                    for cell in row:
                        cell.fill = header_fill
                if row[0].value == 'INDEX':
                    row[0].font = bold_font
                    row[0].fill = header_fill
                if row[0].value == '업무규칙':
                    row[0].font = bold_font
                    row[0].fill = header_fill
            # 파일 저장
            file_path = "./Table_Information.xlsx"
            wb.save(file_path)
        except Exception as e:
            logger.error(f'makeTableList error : {traceback.format_exc()}')